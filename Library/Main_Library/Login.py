#======================== - -       Backend       - - ==============================

#******************  WELCOME TO DIGITAL LIBRARY OF NAVI MUMBAI  ********************

#===================================================================================


# Important Note : ✖️✅➤⚠️
# Please ensure that you had also cloned "Data.txt" file from program link -
# Please change the default path of "Data.txt" to the actual path where you have saved "Data.txt" file -

emp_data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Employee_details.csv"
books_data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx"
data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx"
pre_mem_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Pre_members_info.csv"

#-----------------------------------------------
# Import some important libraries : 
#-----------------------------------------------

import csv
import random
import openpyxl
from colorama import init, Fore, Back, Style
init(autoreset=True)
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

#------------------------------------------------
# Declaring some variables regarding colors and program:
#------------------------------------------------

info_color = Fore.LIGHTYELLOW_EX
text_color = Fore.GREEN
error_color = Fore.RED + Style.BRIGHT
noerror_color = Fore.CYAN + Style.BRIGHT
decor1 = info_color + "*"*55 + "\n"
decor2 = Fore.MAGENTA + "-"*111
decor3 = info_color + "*"*111
decor4 = info_color + "*"*60
filenoterror = f"{Fore.RED + Style.BRIGHT}\n⚠️ Please ensure that you had also cloned 'Library Data' folder from program link ⚠️\n{Fore.RED + Style.BRIGHT}⚠️ Please change the default paths to the actual paths where you have saved Library Data folder, in the variable at line 12 \n"
filecloseerror = f"{decor1}{error_color}Please ensure that you have closed books_data excel file & data excel file ! \n{decor1}"
genres_code = ["MYTH", "CRIMYS", "ROMNC", "BIOGR", "HIS", "NOV", "ECOCIV", "POET", "POLYSC", "MOTV"]
align_centre = Alignment(horizontal='center', vertical='center')

#------------------------------------------------
#  Functions for decorative purposes :
#------------------------------------------------

def underline(text):
    """
    Gives underline to the text.

    Args:
        text: Text to be display in underline.
    
    Returns:
        underlines the text given as parameter. 
    """
    return f"\033[4m{text}\033[0m"

#------------------------------------------------
#  1) Fn of displaying title :
#------------------------------------------------

def login_title():
    """Display attractive title of login portal."""
    a = Back.LIGHTCYAN_EX + "     "
    r = Fore.GREEN + Style.BRIGHT
    b = Fore.BLUE + Style.BRIGHT + Back.LIGHTYELLOW_EX
    re = Fore.RED  + Back.LIGHTMAGENTA_EX
    s = " "*101
    d,t1,T1,t2,t3,t4 = "*"*50,"="*91," "*85,"💻    DIGITAL LIBRARY     💻","OF","🗺️      NAVI MUMBAI       🗺️"
    t5 = "( By MJJ-TECHWORLD )"

    print("\n\n")
    print(Back.LIGHTCYAN_EX + s)
    print(a,r + t1,a, sep = "")
    print(a,r + t1,a, sep = "")
    print(a,r + "|| ",r + T1,r + " ||",a, sep = "")
    print(a,r + "|| ",b + T1,r + " ||",a, sep = "")
    print(a,r + "|| ",b + f"{t2:^83}",r + " ||",a, sep = "")
    print(a,r + "|| ",b + T1,r + " ||",a, sep = "")
    print(a,r + "|| ",b + f"{t3:^85}",r + " ||",a, sep = "")
    print(a,r + "|| ",b + T1,r + " ||",a, sep = "")
    print(a,r + "|| ",b + f"{t4:^87}",r + " ||",a, sep = "")
    print(a,r + "|| ",b + T1,r + " ||",a, sep = "")
    print(a,r + "|| ",b + f"{d:^85}",r + " ||",a, sep = "")
    print(a,r + "|| ",re + f"{t5:^85}",r + " ||",a, sep = "")
    print(a,r + "|| ",r + T1,r + " ||",a, sep = "")
    print(a,r + t1,a, sep = "")
    print(a,r + t1,a, sep = "")
    print(Back.LIGHTCYAN_EX + s, "\n\n")

#------------------------------------------------
#  2) Fn for displaying login options :
#------------------------------------------------

def login_display():
    """
    Take inputs as username & password one by one
    Check whether they are correct or not
    If not, it asks forever, else logged to main interface.
    """

    a = 0 # Initializing
    global username, password
    login_title = "--- LOGIN PORTAL ---"
    print(info_color + "="*100)
    print(Fore.CYAN + f"\n{login_title:^101}\n")
    print(info_color + "*"*100 + "\n\n")

    while True:
        print(decor2)
        username = input(text_color + "Enter Your Username : ").strip()

        if check_username(username) == "yes": 
            while True:
                print(decor2)
                password = input(text_color + "Enter Your Password : ").strip()

                if check_password(username,password) == "yes":
                    a = 1
                    print(decor3)
                    break
            break
    return a

#------------------------------------------------
#  3) Fn for checking username : 
#------------------------------------------------

def check_username(u):
    """
    Check the username whether it is in record or not.
    
    Args:
        u : Username of employee who have access to this program.

    Returns:
        str: "yes" for correct username, "no" for wrong username. 
    """
    try:
        with open(emp_data_path, "r") as f:
            data = csv.reader(f)
            next(data)
            for row in data:
                if (row and row[1] == u):
                    print(f"\n{noerror_color}✅ Username Found \n")
                    return "yes"
                
            print(f"\n{error_color}⚠️  Invalid Username \n")
        
    except FileNotFoundError:
        print(Fore.RED + Style.BRIGHT + "\n⚠️ Please ensure that you had also cloned 'Library Data' folder from program link ⚠️")
        print(Fore.RED + Style.BRIGHT + "⚠️ Please change the default paths to the actual paths where you have saved Library Data folder, in the variable at line 12 \n")

#------------------------------------------------
#  4) Fn to check password :
#------------------------------------------------

def check_password(u,p): 
    """
    Check the password whether it is in record with accordance with its username or not.
    
    Takes username & password as parameters and check accordingly.

    Args:
        u : Username of employee who have access to this program.
        p : Password of employee with registered username as well.

    Returns:
        str: "yes" for correct password, "no" for incorrect password.
    """
    try: 
        with open(emp_data_path, "r") as file:
            data = csv.reader(file)
            next(data)
            for row in data:
                if (row and len(row)>1 and row[1] == u  and row[2] == p):
                    print(f"\n{noerror_color}✅  Logged in successfully! \n")
                    return "yes"

            print(f"\n{error_color}⚠️  Wrong Password \n")
        
    except FileNotFoundError:
        print(Fore.RED + Style.BRIGHT + "\n⚠️ Please ensure that you had also cloned 'Library Data' folder from program link ⚠️")
        print(Fore.RED + Style.BRIGHT + "⚠️ Please change the default paths to the actual paths where you have saved Library Data folder, in the variable at line 12 \n")

#------------------------------------------------
#  ) Fn for taking and checking first name :
#------------------------------------------------
def create_renter_detail():
    """
    Take first name, last name and phone number from users, clear them and check according to need.

    Returns:
        list: 0 or 1 if all details are correct, first_name, last_name, phone_number of renter
    """

    a,b,c = 0,0,0 # Initializing
    while True:
        print(decor2)
        first_name = input(text_color + "Enter first name of renter : ").strip().lower().title()

        if first_name.isalpha() == True:
            print(f"\n{noerror_color}✅  First Name Verified : {first_name}\n")
            a = 1
            break

        else:
            print(f"\n{error_color}⚠️  Please enter valid name\n")           

    if a == 1:
        while True:
            print(decor2)
            last_name = input(text_color + "Enter last name of renter : ").strip().lower().title()

            if last_name.isalpha() == True:
                print(f"\n{noerror_color}✅  Last Name Verified : {last_name}\n")
                b = 1
                break

            else:
                print(f"\n{error_color}⚠️  Please enter valid name\n")

    if b == 1:
        while True:
            print(decor2)
            phone_number = input(text_color + "Enter Phone Number of renter : ")

            if ( phone_number.isdigit() and len(phone_number) == 10 ):
                print(f"\n{noerror_color}✅  Phone Number Verified : {phone_number}\n")
                c = 1
                break
            
            else:
                print(f"\n{error_color}⚠️  Please enter valid phone number\n")

    return [c,first_name,last_name,phone_number]

#------------------------------------------------
#  5) Fn to create title of interface :
#------------------------------------------------

def interface_title():
    """Creates attractice title of main interface after logged in."""
    print(decor2, "\n")
    color0 = Fore.YELLOW
    color1 = Fore.GREEN
    color2 = Fore.CYAN
    color3 = Fore.MAGENTA
    decor = "*"*36
    print(Fore.CYAN + 
    f'''                        \t╔══════════════════════════════════════════════╗
                        \t║                                              ║
                        \t║        {color0}                                      ║
                        \t║        📚  LIBRARY MANAGEMENT SYSTEM  📚     ║
                        \t║        {color1}                                      ║
                        \t║               ( By MJJ-TECHWORLD )           ║
                        \t║                                              ║
                        \t║{color3}     {decor}     ║
                        \t║        {color2}                                      ║
                        \t╚══════════════════════════════════════════════╝''')

    print("\n\n")
    
#------------------------------------------------
#   ) Fn to display actions' options :
#------------------------------------------------

def display_actions():
    """
    Displays actions to be proceed by taking option number as input in str.

    According to option number given by user, it calls to specific functions regarding to it.

    Returns:
        None
    """
    print(f"{info_color}{underline('Available actions')} : \n")
    print(f"\t{noerror_color}1. Search for Unique Code of Book(s) ?")
    print(f"\t{noerror_color}2. Rent Book(s) for rental customers?")
    print(f"\t{noerror_color}3. Rent Book(s) for premium members? \n")
    print(f"\t{noerror_color}4. Return Book(s) of rental customers? \n")
    print(f"\t{noerror_color}5. Return Book(s) of premium members? \n")

    while True:
        print(decor2)
        sel_option = input("Select above option number to proceed further : ").strip()

        if sel_option == "1":
            print(decor2)
            search_uc()
            break

        if sel_option == "2":
            print(decor2)
            rent_by_uc()
            break

        if sel_option == "3":
            print(decor2)
            pre_mem()
            break

        if sel_option == "4":
            print(decor2)
            return_books("Sheet1")
            break

        if sel_option == "5":
            print(decor2)
            return_books("Sheet2")
            break

        else :
            print(f"\n{error_color} ⚠️ Please Enter Correct Option Number (ex. 1 or 2)!\n")

#------------------------------------------------
#   ) Fn to check uc by name of book :
#------------------------------------------------

def check_book_name():
    while True:
        print(decor2)
        book_name = input(text_color + "Enter name of book : ").strip().lower()

        value =  False # Initialising

        wb = load_workbook(books_data_path, data_only=True)
        for s in wb.worksheets:
            for row in s.iter_rows(min_row=2,values_only=True):
                if book_name in str(row[2]).strip().lower():
                    print(decor2)
                    print(f"{row[1]} : {s.title} : {row[2]} : By {row[3]} : Avail {row[7]} : Rs{row[8]}/day")
                    value = True
        
        if value:
            print(decor2)
            break
        else : 
            print(f"\n{error_color}⚠️  Book(s) with this name Not Found ! \n")

#------------------------------------------------
#   ) Fn to check uc by author name of book :
#------------------------------------------------

def check_author_name():
    while True:
        print(decor2)
        author_name = input(text_color + "Enter author name of book : ").strip().lower()

        value =  False # Initialising

        wb = load_workbook(books_data_path, data_only=True)
        for s in wb.worksheets:
            for row in s.iter_rows(min_row=2,values_only=True):
                if author_name in str(row[3]).strip().lower():
                    print(decor2)
                    print(f"{row[1]} : {s.title} : {row[2]} : By {row[3]} : Avail {row[7]} : Rs{row[8]}/day")
                    value = True
            
        if value:
            print(decor2)
            break
        else : 
            print(f"\n{error_color} ⚠️  Book(s) with this name Not Found \n")

#------------------------------------------------
#   ) Fn to check uc by publishing date of book :
#------------------------------------------------

def check_publish_date():
    while True:
        print(decor2)
        publish_date = input(text_color + "Enter publishing date of book in format 'dd-mm-yy' : ").strip().lower()
        print(decor2)

        value = False

        wb = load_workbook(books_data_path, data_only=True)
        for s in wb.worksheets:
            for row in s.iter_rows(min_row=2, values_only=True ):
                if publish_date in str(row[5]).strip().lower():
                    print(decor2)
                    print(f"{row[1]} : {s.title} : {row[2]} : By {row[3]} : Avail {row[7]} : Rs{row[8]}/day")
                    value = True
        
        if value:
            print(decor2)
            break
        else : 
            print(f"\n{error_color} ⚠️  Book(s) with this 'Date Of Publish' Not Found \n")

#------------------------------------------------
#   ) Fn to check uc by publishing date of book :
#------------------------------------------------

def check_genre_book():
    display_genres()
    while True:
        print(decor2)
        genre_books = input(text_color + "Enter genre of desire book from abve table : ").strip().lower()
        print(decor2)

        value = False

        wb = load_workbook(books_data_path, data_only=True)
        for s in wb.worksheets:
            if genre_books in str(s.title).strip().lower():
                for row in s.iter_rows(min_row=2, values_only=True ):
                    print(decor2)
                    print(f"{row[1]} : {s.title} : {row[2]} : By {row[3]} : Avail {row[7]} : Rs{row[8]}/day")
                    value = True
        
        if value:
            print(decor2)
            break
        else : 
            print(f"\n{error_color} ⚠️  Book(s) with this genre Not Found \n")


#------------------------------------------------
#   ) Fn to rent book by searching other factors first then unique code :
#------------------------------------------------

def search_uc():
    print(f"\n{decor4}\n\t{Fore.MAGENTA}   Search Unique Code Of Book(s) \n{decor4}\n")
    while True : 
        print(f"\t{Fore.CYAN}{underline('Ways to find Unique Code (uc) of book(s)')} : ")
        print("\t")
        print("1. By name of book")
        print("2. By author name of book")
        print("3. By publishing date of book")
        print("4. By genre of book (gives list of all books under this genre)\n")

        print(decor2)
        sel_option = input("Select option to proceed further : ").strip()

        if sel_option == "1":
            check_book_name()
            break
        if sel_option == "2":
            check_author_name()
            break
        if sel_option == "3":
            check_publish_date()
            break
        if sel_option == "4":
            check_genre_book()
            break
        else:
            print(f"\n{error_color} ⚠️ Please Enter Correct Option Number (ex. 1 or 2)!\n")

#------------------------------------------------
#   ) Fn to rent book via unique code :
#------------------------------------------------

def rent_by_uc():
    a = 0 # initializing
    s = " "*35
    while True:
        print(decor2)
        unique_code = input(f"{text_color}Enter the unique code of desire book : ")
        print(decor2)

        details = []
        wb = load_workbook(books_data_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2,values_only=False):
                for cell in row:
                   if unique_code == str(cell.value):
                       a = 1
                       details.append(f"{row[1].value} | Name : {row[2].value} |\n{s} Author : {row[3].value} |Language : {row[4].value}")

        if a == 1:
            print(f"\n{noerror_color}✅ Book Found with this unique code {details[0]} \n")
            record_rent_book_detail(unique_code)
            break
        else:
            print(f"\n{error_color}⚠️ Book with this unique code not found !\n")

#------------------------------------------------
#   ) Fn to ask to rent more books or not :
# #------------------------------------------------

def ask_add_more(fn,ln,pn,cl,bl,al,Cl,tl,t,r):
    print(f"{info_color}{underline('Want to rent more books ?')}\n")
    print(f"\t{noerror_color}1. Yes")
    print(f"\t{noerror_color}2. No")

    while True:
        print(decor2)
        sel_option = input("Select above option number to proceed further : ").strip()
        if sel_option == "1":
            print(decor2)
            add_more(fn,ln,pn,cl,bl,al,Cl,tl,t,r)
            break

        if sel_option == "2":
            print(decor2)
            bill_printer(fn,ln,pn,cl,bl,al,Cl,tl,t,r)
            break

        else :
            print(f"\n{error_color} ⚠️ Please Enter Correct Option Number (ex. 1 or 2)!\n")

#------------------------------------------------
#   ) Fn to rent more books : 
#------------------------------------------------

def add_more(fn,ln,pn,cl,bl,al,Cl,tl,t,r="NO"):
    # initializing
    a,b,c = 0,0,0
    s = " "*35
    if r == "NO":
        sn = "Sheet1"
    elif r == "YES":
        sn = "Sheet2"
    while True:
        print(decor2)
        unique_code = input(f"{text_color}Enter the unique code of desire book : ")
        print(decor2)

        details = []
        wb = load_workbook(books_data_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2,values_only=False):
                for cell in row:
                   if unique_code == str(cell.value):
                       a = 1
                       details.append(f"{row[1].value} | Name : {row[2].value} |\n{s} Author : {row[3].value} | Language : {row[4].value}")

        if a == 1:
            print(f"\n{noerror_color}✅ Book Found with this unique code {details[0]} \n")
            while True:
                print(decor2)
                tenure = input(text_color + "Enter borrowing period days : ")
                print(decor2)

                if tenure.isdigit() == True:
                    b = 1
                    break
                else:
                    print(f"{decor1}{error_color}Please enter valid number of days (ex. 3 or 5)\n{decor1}")
                    b = 0

        if b == 1:
            try:
                now_date = datetime.today()
                then_date = now_date + timedelta(days=int(tenure))
                issue_date = now_date.strftime('%d-%m-%Y')
                due_date = then_date.strftime('%d-%m-%Y')
                wb = openpyxl.load_workbook(books_data_path, data_only=True)
                for s in wb.worksheets:
                    for row in s.iter_rows(min_row=2,values_only=False):
                        for cell in row:
                            if unique_code in str(cell.value):
                                c = 1
                                wd = load_workbook(data_path)
                                sheet = wd[sn]
                                print(row[8].value)
                                print(type(row[8].value))
                                print(row[8])
                                sheet.append([f"{fn} {ln}",f"{int(pn)}",row[1].value, row[2].value, row[3].value, row[6].value, row[8].value, issue_date, int(tenure), due_date, int(tenure)*int(row[8].value),"NO"])
                                cl.append(row[1].value)
                                bl.append(row[2].value)
                                al.append(row[3].value)
                                Cl.append(int(row[8].value))
                                tl.append(int(tenure))
                                t.append(int(tenure)*int(row[8].value))
                                for cell in sheet[sheet.max_row]:
                                    cell.alignment = align_centre
                                wd.save(data_path)
                                mod_books(cl,sn,"NO")
                if c == 1:
                    ask_add_more(fn,ln,pn,cl,bl,al,Cl,tl,t,r)

            except FileNotFoundError:
                print(filenoterror)

            except PermissionError:
                print(filecloseerror)
            break
        else:
            print(f"\n{error_color}⚠️ Book with this unique code not found !\n")




#------------------------------------------------
#   ) Fn to record details of book rented : 
#------------------------------------------------

def record_rent_book_detail(uc):
    d,e,code_list,book_list,author_list,charges_list,tenure_list,total = 0,0,[],[],[],[],[],[]
    renter_data = create_renter_detail()
    if renter_data[0] == 1:

        while True:
            print(decor2)
            tenure = input(text_color + "Enter borrowing period days : ")
            print(decor2)

            if tenure.isdigit() == True and tenure != "0":
                d = 1
                break
            else:
                print(f"{decor1}{error_color}Please enter valid number of days (ex. 3 or 5)\n{decor1}")
                d = 0

    if d == 1:
        try:
            now_date = datetime.today()
            then_date = now_date + timedelta(days=int(tenure))
            issue_date = now_date.strftime('%d-%m-%Y')
            due_date = then_date.strftime('%d-%m-%Y')
            wb = openpyxl.load_workbook(books_data_path,data_only=True)
            for s in wb.worksheets:
                for row in s.iter_rows(min_row=2,values_only=False):
                    for cell in row:
                        if uc in str(cell.value):
                            e = 1
                            wd = load_workbook(data_path)
                            sheet = wd["Sheet1"]
                            print(row[8].value)
                            print(type(row[8].value))
                            print(row[8])
                            sheet.append([f"{renter_data[1]} {renter_data[2]}",int(renter_data[3]),row[1].value, row[2].value, row[3].value, row[6].value, row[8].value, issue_date, int(tenure), due_date, int(tenure)*int(row[8].value), "NO"])
                            code_list.append(row[1].value)
                            book_list.append(row[2].value)
                            author_list.append(row[3].value)
                            charges_list.append(int(row[8].value))
                            tenure_list.append(int(tenure))
                            total.append(int(tenure)*int(row[8].value))
                            for cell in sheet[sheet.max_row]:
                                cell.alignment = align_centre
                            wd.save(data_path)
                            mod_books(code_list,"Sheet1","NO")
            if e == 1:
                ask_add_more(renter_data[1], renter_data[2], renter_data[3],code_list,book_list,author_list,charges_list,tenure_list,total,r="NO")


        except FileNotFoundError:
            print(filenoterror)

        except PermissionError:
            print(filecloseerror)

#------------------------------------------------
#   ) Fn to print bill of books rented : 
#------------------------------------------------

def bill_printer(fn,ln,pn,cl,bl,al,Cl,tl,t,r):

    bill_tiltle1 = "📚   DIGITAL LIBRARY OF NAVI MUMBAI   📚"
    bill_tiltle2 = "( By MJJ-TechWorld )"
    l1 = "Name Of Books"
    l2 = "Name Of Authors"
    l3 = "Charges/day"
    l4 = "Tenure"
    l5 = "Total"
    tc = Fore.YELLOW
    l6 = f"{text_color}Date: {tc}{datetime.today().strftime('%d-%m-%Y')}"
    name = f"{fn} {ln}"
    l7 = f"{text_color}Have membership ? {tc}{r}"
    title_text1 = Fore.BLUE
    title_text2 = Fore.RED 
    line = Fore.CYAN
    d = line + "║"


    print(line + "╔",line + "═"*108,line + "╗", sep = "")
    print(d," "*108,d, sep = "")
    print(f"{d}{title_text1}{bill_tiltle1:^106}{d}")
    print(d," "*33,Fore.MAGENTA + "-"*45," "*30,d, sep = "")
    print(d," "*108,d, sep = "")
    print(f"{d}{title_text2}{bill_tiltle2:^108}{d}")
    print(d," "*108,d, sep = "")
    print(d,Fore.MAGENTA + "*"*108,d, sep = "")
    print(d," "*108,d, sep = "")
    print(f"{d} {text_color} Name: {tc}{name:<82}{l6}  {d}")
    print(f"{d} {text_color} Mobile No : {tc}{pn:<71}{l7}  {d}")
    print(d," "*108,d, sep = "")
    print(f"{d} Sr. Unique Code{d}{l1:^36}{d}{l2:^25}{d}{l3:^5}{d}{l4:^6}{d}{l5:^5} {d}")

    for i in range(len(cl)):
        print(f"{d} {i+1:^3}. {cl[i]:^12}{d}{bl[i]:^36}{d}{al[i]:^25}{d}{Cl[i]:^5}{d}{tl[i]:^5}{d}{t[i]:^5} {d}")

    print(d," "*98,d, sep = "")
    print(line + "╚",line + "═"*108,line + "╝", sep = "")

#bill_printer("shvgggggggggggggghg","jvxj","8591095580",["MYTH10001","ECOCIV100001"],["MY DISCOVERY","INDIAN ECONOMY"],["Manish Pandey Dutta", "Ramesh Singh"],["23","34"],["3","4"],["69","136"],"NO")
#ECOCIV10001	Indian Economy	Ramesh Singh	English	01-01-2014	595	8	24
#------------------------------------------------
#   ) Fn to call functions after login : 
#------------------------------------------------

def pre_mem():
    a,b,c,d,cl,bl,al,Cl,tl,t = 0,0,0,0,[],[],[],[],[],[] # Initializing
    while True:
        print(decor2)
        first_name = input(text_color + "Enter first name of member : ").strip().lower().title()

        if first_name.isalpha() == True:
            a = 1
            break

        else:
            print(f"\n{error_color}⚠️  Please enter valid name\n")           

    if a == 1:
        while True:
            print(decor2)
            last_name = input(text_color + "Enter last name of renter : ").strip().lower().title()

            if last_name.isalpha() == True:
                b = 1
                break

            else:
                print(f"\n{error_color}⚠️  Please enter valid name\n")

    if b == 1:
            try:
                with open(pre_mem_path, "r") as f:
                    data = csv.reader(f)
                    next(data)
                    print(Fore.YELLOW + "\nRecords Found from given inputs : ")
                    for row in data:
                        if (row and row[0] == first_name and row[1] == last_name):
                            print(f"{decor2}\n{noerror_color}{row[0]} {row[1]}  :  {row[2]}\n")
                            pn = row[2]
                            c = 1
        
                            if c != 1:
                                print(f"\n{error_color}⚠️ Name not found in data ! \n")
                        
            except FileNotFoundError:
                print(filenoterror)

    if c == 1:
            while True:
                print(decor2)
                phone_number = input(text_color + "Enter Phone Number From above table : ")

                if ( phone_number == pn):

                    otp = random.randint(1000, 9999)
                    print(f"\n{Fore.BLUE}{Style.DIM}Hint : Your OTP is: {otp}")
                    while True:
                        print(decor2)
                        take_otp = input(f"{text_color}Enter otp sent on {phone_number} : ")

                        if take_otp == str(otp):
                            print(f"\n{noerror_color}✅  Member's Details Verified successfully !\n")
                            d = 1
                            break
                        else:
                            print(f"\n{error_color}⚠️ Invalid otp\n")

                    break
                else:
                    print(f"\n{error_color}⚠️ Invalid phone number \n")

    if d == 1:
        add_more(first_name,last_name,phone_number,cl,bl,al,Cl,tl,t,r="YES")

#------------------------------------------------
#   ) Fn to call functions after login : 
#------------------------------------------------

def return_books(m):
    a,b,c,d,nf = 0,0,0,0,[]
    while True:
            print(decor2)
            first_name = input(text_color + "Enter first name of returner : ").strip().lower().title()
    
            if first_name.isalpha() == True:
                a = 1
                break
    
            else:
                print(f"\n{error_color}⚠️  Please enter valid name\n")           
    
    if a == 1:
        while True:
            print(decor2)
            last_name = input(text_color + "Enter last name of returner : ").strip().lower().title()
    
            if last_name.isalpha() == True:
                b = 1
                break

            else:
                print(f"\n{error_color}⚠️  Please enter valid name\n")
    
    if b == 1:
            name = f"{first_name} {last_name}"
            wb = load_workbook(data_path, data_only=True)
            sheet = wb[m]
            for row in sheet.iter_rows(min_row=2,values_only=True):
                if name.strip().lower() == str(row[0]).strip().lower():
                    print(decor2)
                    print(f"{Fore.YELLOW}Name : {row[0]}  |  Phone Number : {row[1]}")
                    pn,c = row[1],1
                else:
                    print("ell")
                    print(row[2])

            wb.save(data_path)
    
    if c == 1:
            while True:
                print(decor2)
                phone_number = input(text_color + "Enter Phone Number From above table : ")

                if ( phone_number == pn):
                    print(decor2)
                    print(f"{info_color}{underline('Important Instructions')}\n")
                    print(f"\t{noerror_color}You can write multiple unique codes separated by commas.")
                    print(f"\t{noerror_color}Example: MYTH10001, NOV10001\n")
                    rbooks = input(text_color + "Enter unique codes of books in given format : ")
                    books = rbooks.replace(" ","")
                    book_list = books.split(",")
                    for book in book_list:
                        wd = load_workbook(data_path, data_only=False)
                        sheet = wd[m]
                        for row in sheet.iter_rows(min_row=2):
                            if book.strip() == str(row[2].value) and str(row[11].value) == "NO":
                                row[11].value = "YES"
                                mod_books(book_list,m,"YES")
                            else:
                                d = 1
                                #print(row[2],row[11])
                                nf.append(book)
    
                        wd.save(data_path)
                    break
                else:
                    print(pn)
                    print(f"\n{error_color}⚠️ Invalid phone number \n")

            print(f"\n{noerror_color}✅  Records Updated successfully !\n")
    if d == 1:
        er =str(nf).replace("[","")
        er = er.replace("]","")
        print(f"\n{error_color}⚠️  Note that book with uc : {er} not found !\n")

#------------------------------------------------
#   ) Fn to call functions after login : 
#------------------------------------------------

def mod_books(l,s,r):
    for book in l:
        wd = load_workbook(books_data_path, data_only=True)

        wd = load_workbook(books_data_path)
        for sheet in wd.worksheets:
            for row in sheet.iter_rows(min_row=2,values_only=False):
                if book == str(row[1].value):
                    if r == "NO":
                        row[7].value = int(row[7].value) - 1
                    elif r == "YES":
                        row[7].value = int(row[7].value) + 1
        wd.save(books_data_path)

#------------------------------------------------
#   ) Fn to call functions after login : 
#------------------------------------------------

def after_login_display():
    interface_title()
    display_actions()

#------------------------------------------------
#   ) Fn for actual login portal
#------------------------------------------------

def start_program():
    login_title()
    if login_display() == 1:
        after_login_display()
        
#------------------------------------------------
#   ) Fn for displaying genres in excel in proper format :
#------------------------------------------------

def display_genres():
    try : 
        global sheet_names
        print(decor2, "\n")

        wb = load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx')
        sheet_names = wb.sheetnames
        for i in sheet_names:
            if sheet_names.index(i) % 2 == 0:
                print(f"   {info_color}{i:<50}|", end = "")
            else:
                print(f"{info_color}{i:>50}")

        print("\n", decor2)

    except FileNotFoundError:
        print(Fore.RED + Style.BRIGHT + "\n⚠️ Please ensure that you had also cloned 'Library Data' folder from program link ⚠️")
        print(Fore.RED + Style.BRIGHT + "⚠️ Please change the default paths to the actual paths where you have saved Library Data folder, in the variable at line 12 \n")


#------------------------------------------------
#    ) Fn to c
#------------------------------------------------

# def check_author_name():
#     while True:
#         print(decor2)
#         author_name = input(text_color + "Enter name of author of book : ").strip().lower()
#         print(decor2)
#         value = False

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if author_name in row[3].strip().lower():
#                 print(decor2)
#                 print(f"{row[1]} : {row[2]} : By {row[3]} : {row[4]} : {row[5]} : Avail {row[7]} : Rs{row[8]}/day")
#                 value = True
            
#         if value:
#             print(decor2)
#             break
#         else : 
#             print(f"{decor1}{error_color}Book(s) with this author name Not Found\n{decor1}")




# def check_publish_date():
#     while True:
#         print(decor2)
#         author_name = input(text_color + "Enter publishing date of book in format 'dd-mm-yy' : ").strip().lower()
#         print(decor2)

#         value = False

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if  author_name in row[5].strip().lower():
#                 print(decor2)
#                 print(f"{row[1]} : {row[2]} : By {row[3]} : Avail {row[7]} : Rs{row[8]}/day")
#                 value = True
            
#         if value:
#             print(decor2)
#             break
#         else : 
#             print(f"{decor1}{error_color}Book(s) with this 'Date Of Publish' Not Found\n{decor1}")


# def call():
#     while True : 
#         print("\tChoose way to find book(s)")
#         print("1. Search by name of book")
#         print("2. Search by name of author of book")
#         print("3. Search by Date of Publish of book")

#         sel_option = input("\nSelect option to proceed further : ")

#         if sel_option == "1":
#             check_book_name()
#             break
#         if sel_option == "2":
#             check_author_name()
#             break
#         if sel_option == "3":
#             check_publish_date()
#             break


def id():
    first_name = input("Enter first name : ")
    last_name = input("Enter last name : ")
    takeid = input("Enter Unique Code : ")
    quantity = input("Enter quantity of books : ")
    t = input("Enter tenure : ")
    from datetime import timedelta, datetime
    now_date = datetime.today()
    then_date = now_date + timedelta(days=int(t))
    issue_date = now_date.strftime('%d-%m-%Y')
    due_date = then_date.strftime('%d-%m-%Y')
    import openpyxl
    wb = openpyxl.load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx', data_only=True)
    a = 0
    details = []
    for s in wb.worksheets:
        for row in s.iter_rows(min_row=2,min_col=2,values_only=False):
            for cell in row:
                if takeid in str(cell.value):
                    a = 1
                    details.append(f"{row[1].value} | Name : {row[2].value} | Author : {row[3].value} | Language : {row[4].value}")
                    ws = openpyxl.load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')
                    sheet = ws.active
                    sheet.append([f"{first_name} {last_name}",row[1].value, row[2].value, row[3].value, row[6].value, int(quantity), row[8].value, issue_date, int(t), due_date, int(quantity)*int(row[8].value)])
                    ws.save(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')



    
    # wb = load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')
    # sheet = wb.active
    # sheet["A2"] = f"{first_name} {last_name}"
    # sheet["B2"] = d[1]
    # sheet["C2"] = d[2]
    # sheet["D2"] = d[3]
    # sheet["E2"] = d[6]
    # sheet["F2"] = q
    # sheet["G2"] = int(d[6])*int(q)
    # wb.save(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')

#==================================================================================================
# Calling functions : 

start_program()


# take_sheet_name()
# first_name = input("Enter first name : ")
# last_name = input("Enter last name : ")
# code = input("Enter unique code : ")
# q = input("Enter quantity of books : ")
# detail = []
# from openpyxl import load_workbook
# wa = load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx')
# sheet = wa["Mythology"]
# for row in sheet.iter_rows(min_row=2,values_only=True):
#     if row[1] == code:
#         detail.append(row)
# d = list(detail[0])
# wa.save(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx')    

# wb = load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')
# sheet = wb.active
# sheet["A2"] = f"{first_name} {last_name}"
# sheet["B2"] = d[1]
# sheet["C2"] = d[2]
# sheet["D2"] = d[3]
# sheet["E2"] = d[6]
# sheet["F2"] = q
# sheet["G2"] = int(d[6])*int(q)
# wb.save(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx')

#------------------------------------------------
#    ) Fn to go to the desire sheet :
#------------------------------------------------

# def take_sheet_name():
#     display_genres()
#     while True :
#         print(decor2)
#         sheet_name = input(text_color + "Enter exact genre of book as given in above table : ").strip()
#         print(decor2)
#         if sheet_name in sheet_names:
#             print(f"{decor1}{noerror_color}Genre Of Book Found\n{decor1}")
#             wb = load_workbook(r'C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx', data_only=True)
#             sheet = wb[sheet_name]
#             call()
#             break
#         else:
#             print(f"{decor1}{error_color}Please enter exact genre of book as given in above table\n{decor1}")
