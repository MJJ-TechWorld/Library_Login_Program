#======================== - -       Backend       - - ==============================

#******************  WELCOME TO DIGITAL LIBRARY OF NAVI MUMBAI  ********************

#===================================================================================


# Important Note : ✖️✅➤⚠️
# Please ensure that you had also cloned "Data.txt" file from program link -
# Please change the default path of "Data.txt" to the actual path where you have saved "Data.txt" file -

emp_data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Employee_details.csv"
books_data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Books_Data.xlsx"
data_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Data.xlsx"
pre_mem_path = r"C:\Users\HP\Desktop\training\Python\Library\Library_Data\Pre_members_info.csv"

#-----------------------------------------------
# Import some important libraries : 
#-----------------------------------------------

import csv
import random
import openpyxl
from colorama import init, Fore, Back, Style
init(autoreset=True)
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Border,Side
from datetime import datetime, timedelta
from Login import underline,display_genres
#------------------------------------------------
# Declaring some variables regarding colors and program:
#------------------------------------------------

info_color = Fore.LIGHTYELLOW_EX
text_color = Fore.GREEN
error_color = Fore.RED + Style.BRIGHT
noerror_color = Fore.CYAN + Style.BRIGHT
decor1 = info_color + "*"*55 + "\n"
decor2 = Fore.MAGENTA + "-"*111
decor3 = info_color + "*"*111
decor4 = info_color + "*"*60
filenoterror = f"{Fore.RED + Style.BRIGHT}\n⚠️ Please ensure that you had also cloned 'Library Data' folder from program link ⚠️\n{Fore.RED + Style.BRIGHT}⚠️ Please change the default paths to the actual paths where you have saved Library Data folder, in the variable at line 12 \n"
filecloseerror = f"{decor1}{error_color}Please ensure that you have closed books_data excel file & data excel file ! \n{decor1}"
genres_code = ["MYTH", "CRIMYS", "ROMNC", "BIOGR", "HIS", "NOV", "ECOCIV", "POET", "POLYSC", "MOTV"]
align_centre = Alignment(horizontal='center', vertical='center')



#------------------------------------------------
#   ) Fn to display actions' options :
#------------------------------------------------

def display_actions():
    """
    Displays actions to be proceed by taking option number as input in str.

    According to option number given by user, it calls to specific functions regarding to it.

    Returns:
        None
    """
    print(f"{info_color}{underline('Available actions')} : \n")
    print(f"\t{noerror_color}1. Search for Unique Code of Book(s) ?")
    print(f"\t{noerror_color}2. Add new genre in data?")
    print(f"\t{noerror_color}3. Rent Book(s) for premium members? \n")
    print(f"\t{noerror_color}4. Return Book(s) of rental customers? \n")
    print(f"\t{noerror_color}5. Return Book(s) of premium members? \n")

    while True:
        print(decor2)
        sel_option = input("Select above option number to proceed further : ").strip()

        if sel_option == "1":
            print(decor2)
        
            break

        if sel_option == "2":
            print(decor2)
            add_new_genre()
            break

        if sel_option == "3":
            print(decor2)

            break

        if sel_option == "4":
            print(decor2)
            
            break

        if sel_option == "5":
            print(decor2)
            
            break

        else :
            print(f"\n{error_color} ⚠️ Please Enter Correct Option Number (ex. 1 or 2)!\n")

#------------------------------------------------
#   ) Fn to display actions' options :
#------------------------------------------------

def add_new_genre():
        try:
            
            print(decor2)
            new_genre = input(text_color + "Enter new genre name : ").strip().upper()
            print(decor2)
            print(Fore.YELLOW + "Add atleast one book's details under this new genre\n")
            new_code = input(text_color + "Enter general unique code for books under this genre (ex. MYTH, NOV) : ").upper()
            print(decor2)
            new_book_name = input(text_color + "Enter proper name of book : ").strip()
            print(decor2)
            new_author_name = input(text_color + "Enter proper name of author of book : ").strip().capitalize()
            print(decor2)
            new_lang = input(text_color + "Enter language in which book is written : ").strip().upper()
            print(decor2)
            new_publ_dt = input(text_color + "Enter publishing date of book in format (dd-mm-yy) : ").strip()
            print(decor2)
            new_mrp = input(text_color + "Enter market price of book (in Rs) : ").strip()
            print(decor2)
            new_qnt = input(text_color + "Enter quantity of copies of this book : ").strip()
            print(decor2)
            new_charge = int((int(new_mrp)*4)/100)
            print(decor2)
                                    
            wb = load_workbook(books_data_path)
            s = wb.create_sheet(new_genre)
            header = ["Sr No","DDC Code","Book Name","Author Name","Language","Published Date", "Market Price (INR)", "Quantities Available", "Charges per day (Rs)"]
            col_width = [7,15,50,28,17,20,18,21,20]
            f = Font(bold=True, underline="single")
            a = Alignment(horizontal="center")
            b = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
            for i in range(len(header)):
                col = i + 1
                c = s.cell(row=1, column=col)
                c.value = header[i]
                c.font = f
                c.alignment = a
                c.border = b
                s.column_dimensions[c.column_letter].width = col_width[i]
            s.append([1,f"{new_code}10001",new_book_name,new_author_name,new_lang,new_publ_dt,int(new_mrp),int(new_qnt),int(new_charge)])
            for cell in s[s.max_row]:
                cell.alignment = align_centre
            print(f"\n{noerror_color}✅ Data updated successfully !\n")
            print(decor2,decor3,sep = "\n")
            wb.save(books_data_path)

        except FileNotFoundError:
            print(filenoterror)
        except PermissionError:
            print(filecloseerror)

#------------------------------------------------
#   ) Fn to display actions' options :
#------------------------------------------------

def add_new_books():
    while True:
        display_genres()
        print(decor2)
        print(Fore.YELLOW + "--- Add New Book Details ---\n")
        new_genre = input(text_color + "Enter genre of this book from above table : ").strip()

        wb = load_workbook(books_data_path, data_only=True)
        req_s = None
        for s in wb.sheetnames:
            if new_genre.strip().lower() in s.strip().lower():
                req_s = s
                break

        if req_s:
            print(decor2)
            new_book_name = input(text_color + "Enter proper name of book : ").strip()
            print(decor2)
            new_author_name = input(text_color + "Enter proper name of author of book : ").strip().capitalize()
            print(decor2)
            new_lang = input(text_color + "Enter language in which book is written : ").strip().upper()
            print(decor2)
            new_publ_dt = input(text_color + "Enter publishing date of book in format (dd-mm-yy) : ").strip()
            print(decor2)
            new_mrp = input(text_color + "Enter market price of book (in Rs) : ").strip()
            print(decor2)
            new_qnt = input(text_color + "Enter quantity of copies of this book : ").strip()
            print(decor2)
            new_charge = int((int(new_mrp)*4)/100)
            print(decor2)

            wd = load_workbook(books_data_path)
            ws = wd[req_s]
            last_row = ws.max_row
            last_sr = ws.cell(row=last_row,column=1).value
            last_uc = ws.cell(row=last_row,column=2).value
            new_sr = int(last_sr) + 1
            text,digt = "",""
            for i in last_uc:
                if i.isdigit() == False:
                    text = text + i
                else:
                    digt = digt + i
            new_digt = int(digt) + 1
            new_uc = text + str(new_digt)
            ws.append([new_sr,new_uc,new_book_name,new_author_name,new_lang,new_publ_dt,int(new_mrp),int(new_qnt),int(new_charge)])
        #s.append([1,f"{new_code}10001",new_book_name,new_author_name,new_lang,new_publ_dt,int(new_mrp),int(new_qnt),int(new_charge)])
            for cell in ws[ws.max_row]:
                cell.alignment = align_centre
            wd.save(books_data_path)
            print(f"\n{noerror_color}✅ Data updated successfully !\n")

        else : 
            print(f"\n{error_color} ⚠️  Book(s) with this genre Not Found \n")



# 
            # wb = load_workbook(books_data_path)
            # for s in wb.sheetnames:
            #     if new_genre in s:
            #         print(decor2)

            # else:
            #     print(f"\n{error_color}⚠️ Genre not found in data \n")


def change_rate():
    wb = load_workbook(books_data_path)
    for sheet in wb.sheetnames:
        s = wb[sheet]
        for row in s.iter_rows(min_row=2,values_only=False):
            if row[6].value is None:
                continue
            row[8].value = int((int(row[6].value) * 3)/100)
    wb.save(books_data_path)         

#add_new_books()